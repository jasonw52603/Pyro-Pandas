import pandas as pd
import numpy as np
import openpyxl as xl
from pandas import ExcelWriter

import config as cfg
import dataFrameFunctions as dfFunc
import os
import traceback


# Init code goes here
print(os.getcwd())
os.chdir('df')
print(os.getcwd())

team_dataframe_list = {}

for file in os.listdir('/home/jasonw/Documents/jwong/Code/Repo/Pyro-Pandas/df'):
    if not 'xlsx' == file[-4:]:  # Only accepts Excel spreadsheets
        continue

    print('----------------------- \n' + file + ':' + '\n')

    wb = xl.load_workbook(filename=file)
    sheetnames = wb.sheetnames
    print(sheetnames)

    # Dataframes
    rawSheet = pd.ExcelFile(file)
    # df0 = pd.read_excel(rawSheet, sheetnames[3])
    # print(df0)

    for sheet_num in range(len(sheetnames)):
        if sheet_num == 0:  # Skips Overview worksheet
            continue

        df0 = pd.read_excel(rawSheet, sheetnames[sheet_num])
        print('Sheet number: ' + str(sheet_num))
        # print(df0)

        df0 = dfFunc.reset_column_names(df0)
        df0 = dfFunc.reset_row_names(df0)

        for rowName in cfg.rowsToIgnore:
            df0 = dfFunc.delete_row(df0, rowName)

        for column in range(3):  # Removes the extra statistics from the scouting app
            df0 = df0[df0.columns[:-1]]

        df0 = dfFunc.reset_row_names(df0)

        # df0 = df0.rename(columns={'dfdf': '1'})
        print('\n \n' + sheetnames[sheet_num] + ' final result:')
        print(df0)
        # team_list[sheetnames[sheet_num]] = df0

        if sheetnames[sheet_num] not in team_dataframe_list:
            team_dataframe_list[sheetnames[sheet_num]] = df0
        else:
            print('found duplicate: ' + sheetnames[sheet_num])
            team_dataframe_list[sheetnames[sheet_num]] = \
                dfFunc.reset_column_names(team_dataframe_list[sheetnames[sheet_num]])

            df0 = dfFunc.delete_column(df0, 0)

            team_dataframe_list[sheetnames[sheet_num]] \
                = team_dataframe_list[sheetnames[sheet_num]].join(df0,
                                                                  lsuffix='_L',
                                                                  rsuffix='_R')

            team_dataframe_list[sheetnames[sheet_num]] = dfFunc.reset_column_names(team_dataframe_list[sheetnames[sheet_num]])
            print(team_dataframe_list[sheetnames[sheet_num]])


print('\n FINAL DATAFRAME LIST OF TEAMS: \n')
leave = False
for team in team_dataframe_list:
    if not leave:
        print(team)
        print('\n' + str(team_dataframe_list[team].to_string()) + '\n')
    if team == '3459 - Team PyroTech':
        leave = True

os.chdir('../outputsheets')
output_path = os.getcwd()
print(os.getcwd())
i = 0
with ExcelWriter(output_path + '/scouting-output.xlsx', mode='w+') as writer:
    for team_data in team_dataframe_list.values():
        team_data.to_excel(excel_writer=writer, sheet_name='FRC Team ' + str(i))

        i += 1
    writer.close()
